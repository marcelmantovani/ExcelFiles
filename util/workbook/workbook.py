import sys
from os import path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook

class WBHandler:
    """
    """

    def __init__(self, in_location:str, inmode:str) -> None:
        
        if (not path.exists(in_location)):
            self._create_new_file(in_location)
    

    def _create_new_file(self, location:str):
        self.wb = Workbook()
        self.wb.save(location)
        self.wb.close()
    
    def _read_workbook(self, in_location:str, inmode='r'):
        self.book = load_workbook(in_location)
        self.writer = pd.ExcelWriter(in_location, engine='openpyxl', mode=inmode)
        self.writer.book = self.book
        self.writer.sheets = dict((ws.title, ws) for ws in self.book.worksheets)
    
    def get_sheet_by_name(self, sheet_name:str) -> pd.DataFrame:
        """
        Loads an Excel sheet of an previously opened excel file
        """

        return pd.read_excel(self.writer, sheet_name=sheet_name)
    
    def save_excel(self, out_location:str, sheet_name:str, df: pd.DataFrame) -> None:
        """
        Saves a Pandas DataFrame given in df to an Excel file name as in out_location,
        using the sheet_name (default is Sheet1).
        Creates a new .xlsx file if it does not already exists, otherwise saves the df to sheet_name
        Parameters
        ----------
        out_location : str
            fully qualified path to the .xlsx file
        df : Pandas.DataFrame
            the df with data to write into Excel
        """

        self._read_workbook(out_location, 'w')
        df.to_excel(self.writer, sheet_name= sheet_name, index=False)
        self.writer.save()

    


